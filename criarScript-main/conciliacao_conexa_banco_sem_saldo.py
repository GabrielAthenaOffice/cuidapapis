import argparse
from collections import defaultdict
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def ler_excel_banco(caminho: str) -> pd.DataFrame:
    """Lê a planilha do Banco e padroniza colunas: Tipo, Protocolo, Data, Valor."""
    df = pd.read_excel(caminho)
    cols = {c.strip(): c for c in df.columns}

    # Campos obrigatórios
    obrig = ["Data", "Valor"]
    faltando = [c for c in obrig if c not in cols]
    if faltando:
        raise ValueError(
            f"Arquivo do Banco precisa ter colunas {faltando}. "
            f"Colunas encontradas: {list(df.columns)}"
        )

    df_out = pd.DataFrame()
    df_out["Tipo"] = df[cols["Tipo"]] if "Tipo" in cols else ""
    df_out["Protocolo"] = df[cols["Protocolo"]] if "Protocolo" in cols else ""
    df_out["Data"] = pd.to_datetime(df[cols["Data"]], errors="coerce", dayfirst=True)
    df_out["Valor"] = pd.to_numeric(
        df[cols["Valor"]].astype(str).str.replace(",", ".", regex=False),
        errors="coerce",
    )
    return df_out[["Tipo", "Protocolo", "Data", "Valor"]]


def ler_excel_conexa(caminho: str) -> pd.DataFrame:
    """Lê a planilha do Conexa e padroniza colunas: Quitação, Valor, Fornecedor (opcional)."""
    df = pd.read_excel(caminho)
    cols = {c.strip(): c for c in df.columns}

    obrig = ["Quitação", "Valor"]
    faltando = [c for c in obrig if c not in cols]
    if faltando:
        raise ValueError(
            f"Arquivo do Conexa precisa ter colunas {faltando}. "
            f"Colunas encontradas: {list(df.columns)}"
        )

    df_out = pd.DataFrame()
    df_out["Quitação"] = pd.to_datetime(df[cols["Quitação"]], errors="coerce", dayfirst=True)
    df_out["Valor"] = pd.to_numeric(
        df[cols["Valor"]].astype(str).str.replace(",", ".", regex=False),
        errors="coerce",
    )
    df_out["Fornecedor"] = df[cols["Fornecedor"]] if "Fornecedor" in cols else ""
    return df_out[["Quitação", "Valor", "Fornecedor"]]


def conciliar_linha_a_linha(banco: pd.DataFrame, erp: pd.DataFrame) -> pd.DataFrame:
    """
    Resultado tem exatamente o mesmo número de linhas do Banco.
    - Se Valor Banco < 0  => tenta conciliar com ERP Valor > 0 (match 1→1 por valor absoluto).
    - Se Valor Banco > 0  => 'Entrada no Banco' (não tenta conciliar).
    - Se sem match        => 'Não conciliado'.
    """
    # Index dos positivos do ERP por valor, para 1→1
    erp_pos = erp[erp["Valor"] > 0].copy()
    index_por_valor = defaultdict(list)
    for idx, row in erp_pos.iterrows():
        index_por_valor[row["Valor"]].append(idx)

    linhas = []
    for _, b in banco.iterrows():
        linha = {
            "Data Banco": b["Data"],
            "Valor Banco": b["Valor"],
            "Tipo Banco": b["Tipo"],
            "Protocolo Banco": b["Protocolo"],
            "Data ERP Quitação": pd.NaT,
            "Valor ERP": pd.NA,
            "Fornecedor ERP": "",
            "Status": ""
        }

        v = b["Valor"]
        if pd.isna(v):
            linha["Status"] = "Zero ou inválido"
        elif v > 0:
            linha["Status"] = "Entrada no Banco"
        else:
            alvo = abs(v)
            candidatos = index_por_valor.get(alvo, [])
            if candidatos:
                idx_erp = candidatos.pop(0)  # consome (garante 1→1)
                e = erp.loc[idx_erp]
                linha["Data ERP Quitação"] = e["Quitação"]
                linha["Valor ERP"] = e["Valor"]
                linha["Fornecedor ERP"] = e.get("Fornecedor", "")
                linha["Status"] = "Conciliado"
            else:
                linha["Status"] = "Não conciliado"

        linhas.append(linha)

    return pd.DataFrame(linhas)


def aplicar_cores(caminho: str, sheet_name: str = "Conciliacao"):
    """Aplica verde p/ positivos e vermelho p/ negativos nas colunas de valores."""
    wb = load_workbook(caminho)
    ws = wb[sheet_name]

    header = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    for col_name in ("Valor Banco", "Valor ERP"):
        col_idx = header.get(col_name)
        if not col_idx:
            continue
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if cell.value > 0:
                        cell.font = Font(color="008000")  # verde
                    elif cell.value < 0:
                        cell.font = Font(color="FF0000")  # vermelho

    # Formatar datas para dd/mm/yyyy
    for col_name in ("Data Banco", "Data ERP Quitação"):
        col_idx = header.get(col_name)
        if not col_idx:
            continue
        letter = get_column_letter(col_idx)
        for cell in ws[f"{letter}2":f"{letter}{ws.max_row}"]:
            cell = cell[0]
            cell.number_format = "dd/mm/yyyy"

    wb.save(caminho)


def main():
    parser = argparse.ArgumentParser(description="Conciliação Banco (base) x Conexa (ERP) – sem saldo diário.")
    parser.add_argument("--banco", required=True, help="Excel do Banco (colunas: Tipo, Protocolo, Data, Valor)")
    parser.add_argument("--erp", required=True, help="Excel do Conexa (colunas: Quitação, Valor, opcional Fornecedor)")
    parser.add_argument("--saida", default="conciliacao_saida.xlsx", help="Arquivo Excel de saída")
    args = parser.parse_args()

    banco = ler_excel_banco(args.banco)
    erp = ler_excel_conexa(args.erp)

    df_conc = conciliar_linha_a_linha(banco, erp)

    out_path = Path(args.saida)
    with pd.ExcelWriter(out_path) as writer:
        df_conc.to_excel(writer, sheet_name="Conciliacao", index=False)

    aplicar_cores(str(out_path))
    print(f"✅ Conciliação gerada em: {out_path.resolve()}")


if __name__ == "__main__":
    main()
